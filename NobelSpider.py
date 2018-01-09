#!/usr/bin/python
# -*- encoding:utf-8 -*-

# pip install requests
# Successfully installed certifi-2017.11.5 chardet-3.0.4 requests-2.18.4 urllib3-1.22

import requests
from openpyxl import *
from bs4 import BeautifulSoup

import sys

reload(sys)
sys.setdefaultencoding('utf8')


def judge_decade(year):
    if year<1910:
        return 0
    elif year<1920:
        return 1
    elif year<1930:
        return 2
    elif year<1940:
        return 3
    elif year<1950:
        return 4
    elif year<1960:
        return 5
    elif year<1970:
        return 6
    elif year<1980:
        return 7
    elif year<1990:
        return 8
    elif year<2000:
        return 9
    elif year<2010:
        return 10
    else:
        return 11


def write_ws(ws_whole, row_id, prize_count_decade):
    for i in range(12):
        ws_whole.cell(row=row_id, column=5+i).value = prize_count_decade[str(i)]


def clear_dict(prize_count_decade):
    for i in range(12):
        prize_count_decade[str(i)]=0


def get_info(web, filename):
    prize_count_decade={'0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0, '10': 0, '11': 0}
    baseurl = 'http://www.nobelprize.org/'
    url = baseurl + web
    response = requests.get(url, timeout=1)
    response.encoding = 'utf-8'  # 将requests强制编码为utf_8
    # print response.encoding  # 查看requests的编码方式

    soup = BeautifulSoup(response.text, 'lxml')
    record_list = soup.find_all('div', {'class': 'by_year'})

    # 写入excel表格
    wb_whole = Workbook()
    ws_whole = wb_whole.get_sheet_by_name("Sheet")
    ws_whole.cell(row=1, column=1).value = 'Institution'
    ws_whole.cell(row=1, column=2).value = 'Location'
    ws_whole.cell(row=1, column=3).value = 'Nation'
    ws_whole.cell(row=1, column=4).value = 'whole_count'
    ws_whole.cell(row=1, column=5).value = '1900_1909'
    ws_whole.cell(row=1, column=6).value = '1910_1919'
    ws_whole.cell(row=1, column=7).value = '1920_1929'
    ws_whole.cell(row=1, column=8).value = '1930_1939'
    ws_whole.cell(row=1, column=9).value = '1940_1949'
    ws_whole.cell(row=1, column=10).value = '1950_1959'
    ws_whole.cell(row=1, column=11).value = '1960_1969'
    ws_whole.cell(row=1, column=12).value = '1970_1979'
    ws_whole.cell(row=1, column=13).value = '1980_1989'
    ws_whole.cell(row=1, column=14).value = '1990_1999'
    ws_whole.cell(row=1, column=15).value = '2000_2009'
    ws_whole.cell(row=1, column=16).value = '2010_2019'

    univ_count = 0
    for i in range(0, len(record_list)):
        prize_cate_year = record_list[i].find('p')
        if not prize_cate_year:  # p is none, then it is a new institution
            if i > 0:
                # store last one
                write_ws(ws_whole, univ_count+1, prize_count_decade)
                # clear to 0
                clear_dict(prize_count_decade)
                print 'write'+str(univ_count)

            rough_data = record_list[i].find('h3').string.strip().split(',')
            l_rough = len(rough_data)
            univ_name = rough_data[0]
            prize_count = rough_data[l_rough-1].split('(')[1].split(')')[0]
            univ_location = ''
            for loc_id in range(1, l_rough-1):
                univ_location = rough_data[loc_id] + ','
            univ_nation = rough_data[l_rough-1].split('(')[0]
            univ_count += 1
            print univ_count
            # write whole
            ws_whole.cell(row=univ_count+1, column=1).value = univ_name
            ws_whole.cell(row=univ_count+1, column=2).value = univ_location
            ws_whole.cell(row=univ_count+1, column=3).value = univ_nation
            ws_whole.cell(row=univ_count+1, column=4).value = prize_count

        else:  # <p> exist
            prize_cate_year_str = prize_cate_year.string.strip()
            l_str = len(prize_cate_year_str)
            prize_split = prize_cate_year_str.split(' ')
            l_split = len(prize_split)
            prize_cate = ''
            pos = prize_cate_year_str.find('in')
            for j in range(pos+3, l_str-5):
                prize_cate += prize_cate_year_str[j]
            prize_year = prize_split[l_split-1]
            decade = judge_decade(int(prize_year))
            prize_count_decade[str(decade)] += 1

    wb_whole.save(filename)
    print 'over'


if __name__ == '__main__':
    get_info('nobel_prizes/lists/universities.html', 'NobelUniversities.xlsx')