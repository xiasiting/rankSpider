#!/usr/bin/python
# -*- encoding:utf-8 -*-

'''
源网址：http://hz.house.ifeng.com/detail/2014_10_28/50087618_1.shtml
项目来源：https://www.zhihu.com/question/26385408
时间：2016-05-19
'''
# pip install requests
# Successfully installed certifi-2017.11.5 chardet-3.0.4 requests-2.18.4 urllib3-1.22

import requests
import xlwt
from openpyxl import *
from bs4 import BeautifulSoup
from pandas import DataFrame
'''
我想去掉面积中的那个㎡，可是报了
UnicodeDecodeError: 'ascii' codec can't decode byte 0xe3 in position 0: ordinal not in range(128)
这个错误，所以就加上了下面三句，解决
'''
import sys
reload(sys)
sys.setdefaultencoding('utf8')

def get_info(web, filename):
    worldrank=[]
    institution=[]
    location=[]
    nationalrank=[]
    totalscore=[]
    alumni=[]
    award=[]
    hici=[]
    n_s=[]
    pub=[]
    pcp=[]

    baseurl = 'http://shanghairanking.com/'
    url = baseurl+web+'.html'
    response=requests.get(url)
    response.encoding = 'utf-8' #将requests强制编码为utf_8
    # print response.encoding   查看requests的编码方式

    soup = BeautifulSoup(response.text, 'lxml')
    arcicle = soup.find('div', {'id':'rankingarea'})

    tr = arcicle.find_all('tr')
    # 写入excel表格
    wb = Workbook()
    ws1 = wb.get_sheet_by_name("Sheet")
    for i in range(1, len(tr)):
        td = tr[i].find_all('td')
        
        worldrank.append(td[0].string.strip())
        ws1.cell(row=i + 1, column=1).value = td[0].string.strip()

        institution.append(td[1].string.strip())
        ws1.cell(row=i + 1, column=2).value = td[1].string.strip()

        imgname = td[2].find_all('img')
        imgsrc = imgname[0].get('src')
        locationname = imgsrc.split('/')[2].split('.')[0]
        location.append(locationname)
        ws1.cell(row=i + 1, column=3).value = locationname


        nranknumber = td[3].find_all('div')
        nationalrank.append(nranknumber[0].string.strip())
        ws1.cell(row=i + 1, column=4).value = nranknumber[0].string.strip()

        t_score = td[4].find_all('div')
        # 有可能为空
        if t_score[0].string:
            totalscore.append(t_score[0].string.strip())
            ws1.cell(row=i + 1, column=5).value = t_score[0].string.strip()

        alumni_score = td[5].find_all('div')
        alumni.append(alumni_score[0].string.strip())
        ws1.cell(row=i + 1, column=6).value = alumni_score[0].string.strip()

        award_score = td[6].find_all('div')
        award.append(award_score[0].string.strip())
        ws1.cell(row=i + 1, column=7).value = award_score[0].string.strip()

        hici_score = td[7].find_all('div')
        hici.append(hici_score[0].string.strip())
        ws1.cell(row=i + 1, column=8).value = hici_score[0].string.strip()

        n_s_score = td[8].find_all('div')
        if n_s_score[0].string:
            n_s.append(n_s_score[0].string.strip())
            ws1.cell(row=i + 1, column=9).value = n_s_score[0].string.strip()

        pub_score = td[9].find_all('div')
        pub.append(pub_score[0].string.strip())
        ws1.cell(row=i + 1, column=10).value = pub_score[0].string.strip()

        pcp_score = td[10].find_all('div')
        pcp.append(pcp_score[0].string.strip())
        ws1.cell(row=i + 1, column=11).value = pcp_score[0].string.strip()

    ws1.cell(row=1, column=1).value = 'worldrank'
    ws1.cell(row=1, column=2).value = 'institution'
    ws1.cell(row=1, column=3).value = 'location'
    ws1.cell(row=1, column=4).value = 'nationalrank'
    ws1.cell(row=1, column=5).value = 'totalscore'
    ws1.cell(row=1, column=6).value = 'alumni_score'
    ws1.cell(row=1, column=7).value = 'award_score'
    ws1.cell(row=1, column=8).value = 'hici_score'
    ws1.cell(row=1, column=9).value = 'n&s_score'
    ws1.cell(row=1, column=10).value = 'pub_score'
    ws1.cell(row=1, column=11).value = 'pcp_score'
    wb.save(filename)

    df = DataFrame(worldrank, columns=['worldrank'])
    df['institution'] = DataFrame(institution)
    df['nationalrank'] = DataFrame(nationalrank)
    df['totalscore'] = DataFrame(totalscore)
    df['alumni_score'] = DataFrame(alumni)
    df['award_score'] = DataFrame(award)
    df['hici_score'] = DataFrame(hici)
    df['n&s_score'] = DataFrame(n_s)
    df['pub_score'] = DataFrame(pub)
    df['pcp_score'] = DataFrame(pcp)

    return df


if __name__ == '__main__':

    page2017 = get_info('ARWU2017', 'shanghai2017.xlsx')
    page2016 = get_info('ARWU2016', 'shanghai2016.xlsx')